from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
import logging

app = Flask(__name__)

# Configure the logger
logging.basicConfig(filename='error.log', level=logging.ERROR)
logger = logging.getLogger()

@app.route('/crawl', methods=['POST'])
def crawl():
    try:
        # Code for web crawling and data extraction
        
        # If an error occurs, raise an exception
        raise Exception("An error occurred during the crawling process.")
        
        # Return the extracted data
        
    except Exception as e:
        # Log the error
        logger.error(str(e))
        
        # Return an error response to the frontend
        return jsonify({'error': 'An error occurred during the crawling process.'}), 500
# ...

@app.route('/process', methods=['POST'])
def process():
    keyword = request.json['keyword']
    
    # Perform web scraping and extract contact information based on the keyword
    extracted_data = [
        {'name': 'John Doe', 'email': 'john.doe@example.com'},
        {'name': 'Jane Smith', 'email': 'jane.smith@example.com'}
    ]
    
    # Update the Excel sheet with the extracted contact information
    workbook: Workbook = load_workbook('contacts.xlsx')
    worksheet = workbook.active

    # Write the headers if the Excel sheet is empty
    if not worksheet['A1'].value:
        worksheet['A1'] = 'Name'
        worksheet['B1'] = 'Email'

    # Write the data
    row_count = worksheet.max_row + 1
    for index, data in enumerate(extracted_data, start=row_count):
        worksheet[f'A{index}'] = data['name']
        worksheet[f'B{index}'] = data['email']

    # Save the changes and close the workbook
    workbook.save('contacts.xlsx')
    workbook.close()
    
    # Return the extracted data as a response to the frontend
    result = {
        'keyword': keyword,
        'contact_info': extracted_data
    }
    
    return jsonify(result)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000)




app = Flask(__name__)

@app.route('/')
def index():
    return app.send_static_file('index.html')

@app.route('/process', methods=['POST'])
def process():
    keyword = request.json['keyword']
    
    # Perform web scraping and extract contact information based on the keyword
    # Store the extracted data in a structured format (e.g., JSON or CSV)
    
    # Return the extracted data as a response to the frontend
    result = {
        'keyword': keyword,
        'contact_info': contactData  # Replace 'extracted_data' with the actual extracted data
    }
    
    return jsonify(result)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000)
